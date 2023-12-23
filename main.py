from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PilImage

def insertimg2excel(imgPath, excelPath,row_num,col_num):

    imgsize = (720 / 4, 1280 / 4)  # 设置一个图像缩小的比例
    wb = load_workbook(excelPath)
    ws = wb.active
    ws[row_num + str(col_num)].value = None
    ws.column_dimensions[row_num].width = imgsize[0] * 0.14  # 修改列A的宽

    img = Image(imgPath)  # 缩放图片
    img.width, img.height = imgsize

    ws.add_image(img, row_num + str(col_num))  # 图片 插入 A1 的位置上
    ws.row_dimensions[col_num].height = imgsize[1] * 0.78  # 修改列第1行的宽

    wb.save('out1.xlsx')  # 新的结果保存输出


if __name__ == '__main__':
    imgPath = 'a.jpg'
    excelPath = 'test.xlsx'
    insertimg2excel(imgPath, excelPath,'C',5)
    insertimg2excel(imgPath, excelPath, 'C', 4)