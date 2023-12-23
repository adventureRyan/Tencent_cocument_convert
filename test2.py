from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PilImage

wb = load_workbook('test.xlsx')

ws = wb.active

column_c = ws['C']

for cell in column_c:
    print(cell.value)
