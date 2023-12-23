import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image


def resize_image(img_path, new_width=600):
    image = Image.open(img_path)
    size = image.size
    img_ratio = size[0] / size[1]
    new_image = image.resize((new_width, int(new_width / img_ratio)))
    new_image.save(img_path)

def insert_images_to_excel(excel_path):
    # Load the workbook
    wb = load_workbook(excel_path)
    ws = wb.active

    # Iterate over cells in column C
    for cell in ws['C']:
        # Check if a file with the same name as the cell value exists in the img directory
        img_path = os.path.join('img',str(cell.value))
        cnt = 1
        if os.path.exists(img_path):
            # Load the image and resize it
            print("h")
            resize_image(img_path, 100)
            img = Image(img_path)
            image = Image.open(img_path)
            imgsize = image.size
            ws.column_dimensions['C'].width = imgsize[0] * 0.14  # 修改列A的宽

            ws.add_image(img, 'C' + str(cnt))
            cnt += 1
            ws.row_dimensions[cnt].height = imgsize[1] * 0.78
    wb.save('out.xlsx')  # 新的结果保存输出


    # Save the workbook
    wb.save('output.xlsx')

if __name__ == '__main__':
    excel_path = 'test.xlsx'
    img_dir = 'img'
    insert_images_to_excel(excel_path, img_dir)