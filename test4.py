import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def resize_image(img_path, new_width=600):
    image = Image.open(img_path)
    size = image.size
    img_ratio = size[0] / size[1]
    new_image = image.resize((new_width, int(new_width / img_ratio)))
    new_image.save(img_path)

def insertimg2excel(wb,ws,imgPath, excelPath,row_num,col_num):

    imgsize = (720 / 4, 1280 / 4)  # 设置一个图像缩小的比例
    ws[row_num + str(col_num)].value = None
    ws.column_dimensions[row_num].width = imgsize[0] * 0.14  # 修改列A的宽

    img = Image(imgPath)  # 缩放图片
    img.width, img.height = imgsize

    ws.add_image(img, row_num + str(col_num))  # 图片 插入 A1 的位置上
    ws.row_dimensions[col_num].height = imgsize[1] * 0.78  # 修改列第1行的宽


def insert_images_to_excel(excel_path):
    # Load the workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    cnt = 1
    for cell in ws['C']:
        img_path = os.path.join('img',str(cell.value))

        if os.path.exists(img_path):
            print("H")
            insertimg2excel(wb,ws,img_path, excel_path, 'C', cnt)
        cnt += 1

    wb.save('out.xlsx')  # 新的结果保存输出

if __name__ == '__main__':
    excel_path = '1.xlsx'
    img_dir = '图片'
    insert_images_to_excel(excel_path)